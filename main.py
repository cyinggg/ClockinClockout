import telebot
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime

# Load environment variables
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise Exception("BOT_TOKEN not found. Please check .env file!")

bot = telebot.TeleBot(BOT_TOKEN)

# Excel files
STUDENTS_FILE = "studentCoach.xlsx"
STAFF_FILE = "staff.xlsx"
LOG_FILE = "log.xlsx"

# Telegram group/channel ID for supervisor notifications
SUPERVISOR_CHAT_ID = -

# ======================
# Helper Functions
# ======================

def get_student_name(student_id):
    """Return student name from studentCoach.xlsx"""
    wb = load_workbook(STUDENTS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_id = str(row[0]).split(".")[0]
        if str(student_id).strip() == cell_id:
            return row[1]
    return None


def is_valid_staff(staff_name):
    """Check if staff name exists in staff.xlsx"""
    wb = load_workbook(STAFF_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().lower() == staff_name.lower():
            return True
    return False


def log_action(student_id, action):
    """Log clock in/out to Excel"""
    name = get_student_name(student_id)
    if not name:
        return False

    wb = load_workbook(LOG_FILE)
    ws = wb.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([timestamp, student_id, name, action, "Pending", "", ""])
    wb.save(LOG_FILE)

    bot.send_message(
        SUPERVISOR_CHAT_ID,
        f"📌 *{action} request pending verification*\n👤 Student: {name} ({student_id})\n🕒 Time: {timestamp}",
        parse_mode="Markdown"
    )
    return True


def get_pending_records():
    """Return a list of (index, student_name, action, timestamp) for pending verification"""
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    records = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[4]).strip().lower() == "pending":  # Verification Status column
            records.append((idx, row[2], row[3], row[0]))  # row[2]=name, row[3]=action, row[0]=timestamp
    return records


def update_verification(rows_to_verify, staff_name, signature):
    """Mark selected rows as verified with staff name, signature, and verified timestamp"""
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    verified_students = []

    for row_num in rows_to_verify:
        ws.cell(row=row_num, column=5, value="Verified")        # Status
        ws.cell(row=row_num, column=6, value=signature)         # Signature
        ws.cell(row=row_num, column=7, value=timestamp)         # Verified Timestamp
        verified_students.append(ws.cell(row=row_num, column=3).value)  # Student name

    wb.save(LOG_FILE)

    # Send group notification
    student_list = "\n".join([f"- {name}" for name in verified_students])
    bot.send_message(
        SUPERVISOR_CHAT_ID,
        f"✅ *Verification Completed*\n🗓 Date: {timestamp}\n👨‍🏫 Verified by: {staff_name}\n📋 Students Verified:\n{student_list}",
        parse_mode="Markdown"
    )


# ======================
# Telegram Bot Commands
# ======================

@bot.message_handler(commands=['start'])
def start(message):
    bot.reply_to(message, "👋 Welcome! Use /clockin or /clockout to record your shift.\nStaff can use /verify to confirm student records.")


@bot.message_handler(commands=['clockin'])
def clock_in(message):
    msg = bot.reply_to(message, "Please enter your Student ID to Clock In:")
    bot.register_next_step_handler(msg, process_clock_in)

def process_clock_in(message):
    student_id = message.text.strip()
    if log_action(student_id, "Clock In"):
        bot.reply_to(message, f"✅ Clock In recorded for {student_id}")
    else:
        bot.reply_to(message, "❌ Student ID not found. Please try again.")


@bot.message_handler(commands=['clockout'])
def clock_out(message):
    msg = bot.reply_to(message, "Please enter your Student ID to Clock Out:")
    bot.register_next_step_handler(msg, process_clock_out)

def process_clock_out(message):
    student_id = message.text.strip()
    if log_action(student_id, "Clock Out"):
        bot.reply_to(message, f"✅ Clock Out recorded for {student_id}")
    else:
        bot.reply_to(message, "❌ Student ID not found. Please try again.")


@bot.message_handler(commands=['verify'])
def verify_records(message):
    pending = get_pending_records()
    if not pending:
        bot.reply_to(message, "✅ No pending records for verification.")
        return

    msg_text = "📋 *Pending Clock In/Out Records:*\n"
    for idx, (row_num, name, action, time) in enumerate(pending, start=1):
        msg_text += f"{idx}. {name} — {action} at {time}\n"
    msg_text += "\nEnter the record numbers to verify (e.g. `1,3,5` or `all`):"

    bot.send_message(message.chat.id, msg_text, parse_mode="Markdown")
    bot.register_next_step_handler(message, lambda m: select_records_to_verify(m, pending))


def select_records_to_verify(message, pending):
    selection = message.text.strip().lower()
    if selection == "all":
        rows_to_verify = [r[0] for r in pending]
    else:
        try:
            indexes = [int(i) for i in selection.split(",")]
            rows_to_verify = [pending[i-1][0] for i in indexes]
        except:
            bot.reply_to(message, "❌ Invalid selection. Please try again.")
            return

    msg = bot.reply_to(message, "Please enter your *staff name* for verification:", parse_mode="Markdown")
    bot.register_next_step_handler(msg, lambda m: verify_staff_identity(m, rows_to_verify))


def verify_staff_identity(message, rows_to_verify):
    staff_name = message.text.strip()
    if not is_valid_staff(staff_name):
        bot.reply_to(message, "❌ Staff name not found in records.")
        return

    msg = bot.reply_to(message, "✍️ Please enter your *signature* to confirm verification:", parse_mode="Markdown")
    bot.register_next_step_handler(msg, lambda m: finalize_verification(m, staff_name, rows_to_verify))


def finalize_verification(message, staff_name, rows_to_verify):
    signature = message.text.strip()
    update_verification(rows_to_verify, staff_name, signature)
    bot.reply_to(message, f"✅ Verified successfully by {staff_name}.\nSignature saved for selected records.")


@bot.message_handler(commands=['getid'])
def get_id(message):
    bot.reply_to(message, f"💬 Chat ID: `{message.chat.id}`", parse_mode="Markdown")


# ======================
# Run Bot
# ======================
print("🤖 WorkBot is running...")
bot.infinity_polling()
